import os
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from docx import Document
from io import BytesIO

def convert_to_txt(file): #DataFrame을 |으로 구분된 CSV 파일로 변환
    df = pd.read_excel(file)
    txt_output = df.to_csv(index=False, sep='|')
    return txt_output

def convert_to_json(file): #DataFrame을 레코드별로 JSON 형식으로 변환
    df = pd.read_excel(file)
    json_output = df.to_json(orient='records')
    return json_output

def convert_to_log(file): #시트의 각 행을 |으로 구분된 텍스트 형식으로 변환
    wb = load_workbook(file)
    ws = wb.active
    log_content = ''
    for row in ws.iter_rows(values_only=True): 
        log_content += '|'.join(str(cell) for cell in row) + '\n'
    return log_content

def convert_to_docx(file): #각 행을 문단으로 추가하여 DOCX 형식으로 변환
    wb = load_workbook(file)
    ws = wb.active
    doc = Document()
    for row in ws.iter_rows(values_only=True): 
        doc.add_paragraph('|'.join(str(cell) for cell in row))
    return doc


st.title('엑셀 파일 변환')

uploaded_file = st.file_uploader('엑셀 파일을 업로드 해주세요', type=['xlsx'])

if uploaded_file:
    st.write('파일 업로드 성공')
    uploaded_filename = os.path.splitext(uploaded_file.name)[0]

    st.download_button(label='Download TXT', data=convert_to_txt(uploaded_file), file_name=f'{uploaded_filename}.txt')
    st.download_button(label='Download JSON', data=convert_to_json(uploaded_file), file_name=f'{uploaded_filename}.json')
    st.download_button(label='Download LOG', data=convert_to_log(uploaded_file), file_name=f'{uploaded_filename}.log')

    docx_file = convert_to_docx(uploaded_file)
    docx_content = BytesIO()
    docx_file.save(docx_content) 
    docx_content.seek(0)
    st.download_button(label='Download DOCX', data=docx_content, file_name=f'{uploaded_filename}.docx')