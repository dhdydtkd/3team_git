import requests
from collections import Counter
import time 
import re
import openpyxl
from datetime import datetime
from faker import Faker
import random


def get_criminalip_info(ip_address):
    url = f"https://api.criminalip.io/v1/feature/ip/is_safe_dns_server?ip={ip_address}"
    payload={}
    headers = {
        "x-api-key": "GUr2z0Om1FHnKt5E1Oh9yTngidRCZg1myYOESHnpdUU7hl6majhCSaj7L95v"
    }
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        return response.json()
    else:
        print(f'criminalip에서 오류가 발생했습니다.')
    

def get_file_ip(log_file):
    ip_list = []
    with open(log_file, 'r') as f:
        for line in f:
            #match = re.findall(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", line)
            match = re.search(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", line)
            if match:
                ip_list.append(match[0])

    return ip_list


if __name__ == "__main__":
    create_member_count = 1000
    log_file = 'access.log'
    ip_list = get_file_ip(log_file)
    ip_counter = Counter(ip_list)
    ips = ip_counter.most_common(create_member_count)

    workbook = openpyxl.Workbook()
    # 랜덤 마스킹
    fake = Faker('ko_KR')
    for index,ip_address in enumerate(ips):
        result = get_criminalip_info(ip_address[0])
        if result and result['ip'] :
            worksheet = workbook.active
            worksheet['A1'] = "이름"
            worksheet['B1'] = "핸드폰"
            worksheet['C1'] = "이메일"
            worksheet['D1'] = "주소"
            worksheet['E1'] = "상태"
            worksheet['F1'] = "ip"
            worksheet['G1'] = "is_safe_dns_server"
            worksheet['H1'] = "is_malicious_ip"
            worksheet['I1'] = "opened_unusual_port"
            worksheet['J1'] = "has_dns_vulnerability"
            worksheet['K1'] = "is_unpublic_dns"
            worksheet['L1'] = "create_dt"
            worksheet.cell(row=index+2, column=1, value=fake.name())
            if random.randrange(1,30)>=29 : 
                worksheet.cell(row=index+2, column=2, value=fake.phone_number())
            else: 
                phone_1, phone_2, phone_3 = fake.phone_number().split('-')
                masked_phone = phone_1 + '-' + '*' * len(phone_2) + '-' + phone_3
                worksheet.cell(row=index+2, column=2, value=masked_phone)
            
            if random.randrange(1,30)>=29 : 
                worksheet.cell(row=index+2, column=3, value=fake.email())
            else: 
                username, domain = fake.email().split('@')
                masked_email = '*' * len(username) + '@' + domain
                worksheet.cell(row=index+2, column=3, value=masked_email)
                
            worksheet.cell(row=index+2, column=4, value=fake.address())
            worksheet.cell(row=index+2, column=5, value=result['status'])
            
            worksheet.cell(row=index+2, column=6, value=result['ip'])
            if random.randrange(1,10)>=9 : 
                worksheet.cell(row=index+2, column=7, value=False)
                flag=True
                while(flag) : 
                    if random.randrange(1,10)>=9 : 
                        worksheet.cell(row=index+2, column=8, value=True)
                        flag = False
                    else: 
                        worksheet.cell(row=index+2, column=8, value=False)
                    if random.randrange(1,10)>=9 : 
                        worksheet.cell(row=index+2, column=9, value=True)
                        flag = False
                    else: 
                        worksheet.cell(row=index+2, column=9, value=False)
                    if random.randrange(1,10)>=9 : 
                        worksheet.cell(row=index+2, column=10, value=True)
                        flag = False
                    else: 
                        worksheet.cell(row=index+2, column=10, value=False)
                    if random.randrange(1,10)>=9 : 
                        worksheet.cell(row=index+2, column=11, value=True)
                        flag = False
                    else: 
                        worksheet.cell(row=index+2, column=11, value=False)
            else : 
                worksheet.cell(row=index+2, column=7, value=True)
                worksheet.cell(row=index+2, column=8, value=False)
                worksheet.cell(row=index+2, column=9, value=False)
                worksheet.cell(row=index+2, column=10, value=False)
                worksheet.cell(row=index+2, column=11, value=False)
            worksheet.cell(row=index+2, column=12, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            time.sleep(1)
    workbook.save(f'member_data/{datetime.now().strftime('%Y-%m-%d')}_insert_member.xlsx')